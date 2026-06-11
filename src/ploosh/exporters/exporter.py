# pylint: disable=R0903,W0613
"""Test result exporter"""
from datetime import datetime
import os

import pandas as pd
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


class Exporter:
    """Test result exporter"""
    name = None  # Name of the exporter
    output_path = None  # Output path for the exported results

    @staticmethod
    def date_to_string(data):
        """Convert datetime to string in ISO 8601 format"""
        if not isinstance(data, datetime):
            return None

        return data.strftime("%Y-%m-%dT%H:%M:%SZ")

    def export(self, cases: dict, execution_id: str):
        """Export test case results to the destination"""
        return None


    def export_gap_file(self, detail_file_path: str, df_compare_gap):
        """Export comparison gap to a styled Excel file."""

        os.makedirs(os.path.dirname(detail_file_path), exist_ok=True)

        df = df_compare_gap.copy()

        # Remove columns and rows that are entirely empty to avoid clutter in the output file.
        df = df.dropna(axis=1, how="all").dropna(axis=0, how="all")

        df.index.name = "row"

        with pd.ExcelWriter(detail_file_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="gap", merge_cells=True)
            worksheet = writer.sheets["gap"]

            header_column_fill = PatternFill("solid", fgColor="244062")
            header_type_fill = PatternFill("solid", fgColor="366092")

            diff_fill = PatternFill("solid", fgColor="F6C9CE")
            diff_font = Font(color="9C0006")
            match_fill = PatternFill("solid", fgColor="C6EFCE")
            match_font = Font(color="006100")
            header_font = Font(bold=True, color="FFFFFF")
            index_fill = PatternFill("solid", fgColor="DCE6F1")
            index_font = Font(bold=True)
            center = Alignment(horizontal="center", vertical="center")
            thin = Side(style="thin", color="BFBFBF")

            header_rows = df.columns.nlevels  # 2 levels: column name / expected-source
            value_columns = len(df.columns)

            # Remove the default header row added by pandas and keep only the multi-level column headers.
            worksheet.delete_rows(header_rows + 1)

            data_start_row = header_rows + 1
            data_end_row = data_start_row + len(df) - 1
            last_column = value_columns + 1  # +1 for the index column

            # Merge the index header cell and style it.
            worksheet.merge_cells(start_row=1, start_column=1, end_row=header_rows, end_column=1)
            index_header = worksheet.cell(row=1, column=1, value="row")
            index_header.fill = header_column_fill
            index_header.font = header_font
            index_header.alignment = center

            # Style the column headers.
            for row in range(1, header_rows + 1):
                for column in range(2, last_column + 1):
                    cell = worksheet.cell(row=row, column=column)
                    if isinstance(cell, MergedCell):
                        # Non-anchor merged cells are read-only; the anchor carries the style.
                        continue
                    cell.fill = header_type_fill if row == header_rows else header_column_fill
                    cell.font = header_font
                    cell.alignment = center

            # Style the data area and apply the conditional coloring.
            group_names = df.columns.get_level_values(0)
            for offset in range(len(df)):
                row = data_start_row + offset

                index_cell = worksheet.cell(row=row, column=1)
                index_cell.fill = index_fill
                index_cell.font = index_font
                index_cell.alignment = center

                for group in group_names.unique():
                    positions = [i for i, name in enumerate(group_names) if name == group]
                    has_gap = any(not pd.isna(df.iat[offset, position]) for position in positions)
                    fill = diff_fill if has_gap else match_fill
                    font = diff_font if has_gap else match_font
                    for position in positions:
                        cell = worksheet.cell(row=row, column=position + 2)
                        cell.fill = fill
                        cell.font = font
                        cell.alignment = center

                        cell.border = Border(
                            left=thin,
                            right=Side(style="thin", color="000000") if position == positions[-1] else thin,
                            top=thin,
                            bottom=thin
                        )

            # Auto-size columns based on their content.
            for column in range(1, last_column + 1):
                letter = get_column_letter(column)
                lengths = [
                    len(str(worksheet.cell(row=row, column=column).value))
                    for row in range(1, data_end_row + 1)
                    if worksheet.cell(row=row, column=column).value is not None
                ]
                worksheet.column_dimensions[letter].width = (max(lengths) if lengths else 5) + 2

